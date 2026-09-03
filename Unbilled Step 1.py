import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path
import re
from datetime import datetime

from billing_rules import (
    parse_weekday_from_token,
    is_non_billable_service_for_weekday,
    is_professional_claim_type,
    is_iop_service,
    is_php_service,
    is_cathy_payer,
    is_cathy_all_payer,
    is_aetna_payer,
    _is_drug_screen as is_drug_screen,
    CATHY_PAYERS,
    CATHY_ALL_PAYERS,
    ROSANNA_PROFESSIONAL_CAP,
)

stubs_required = False  # no Streamlit in the standalone script; keep for parity if needed

def extract_date_from_filename(filename):
    """Extract MMDDYYYY date from filename"""
    match = re.search(r'(\d{8})', filename)
    if match:
        return match.group(1)
    return None

def is_wm_program_level(cell_value) -> bool:
    """Return True if the Program Level cell contains 'WM'."""
    return "WM" in str(cell_value or "").upper()

def get_save_folder(workbook_path, date_token):
    """Build save path: ...\\Unbilled Reports\\YYYY\\MM - Month YYYY\\MMDDYYYY"""
    wb_path = Path(workbook_path)

    # Find "Unbilled Reports" in the path
    parts = wb_path.parts
    try:
        idx = [i for i, p in enumerate(parts) if 'Unbilled Reports' in p][0]
        base = Path(*parts[:idx+1])
    except IndexError:
        raise ValueError("Workbook must be saved in a folder containing 'Unbilled Reports'")

    # Parse date token: MMDDYYYY
    month = date_token[:2]
    day = date_token[2:4]
    year = date_token[4:]

    dt = datetime.strptime(date_token, '%m%d%Y')
    month_name = dt.strftime('%B')

    # Build: YYYY\\MM - Month YYYY\\MMDDYYYY
    save_path = base / year / f"{month} - {month_name} {year}" / date_token
    save_path.mkdir(parents=True, exist_ok=True)

    return save_path

def step_1_extract_invalid(ws):
    """Insert Staff/Status, extract Invalid rows to new sheet, delete from main"""

    # Insert Staff/Status column if needed
    if ws.cell(1, 1).value != "Staff/Status":
        ws.insert_cols(1)
        ws.cell(1, 1).value = "Staff/Status"

    # Find Billable Status column
    billable_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(1, col).value == "Billable Status":
            billable_col = col
            break

    if not billable_col:
        raise ValueError("Billable Status column not found")

    # Find all invalid rows
    invalid_rows = []
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, billable_col).value == "Invalid for Billing":
            invalid_rows.append(row)

    if not invalid_rows:
        print("No invalid rows found")
        return None

    # Create Invalid sheet
    wb = ws.parent
    if "Invalid" in wb.sheetnames:
        del wb["Invalid"]

    ws_invalid = wb.create_sheet("Invalid")

    # Copy header
    for col in range(1, ws.max_column + 1):
        ws_invalid.cell(1, col).value = ws.cell(1, col).value

    # Copy invalid rows
    for idx, row_num in enumerate(invalid_rows, start=2):
        for col in range(1, ws.max_column + 1):
            ws_invalid.cell(idx, col).value = ws.cell(row_num, col).value

    # Delete invalid rows (backwards to preserve row numbers)
    for row_num in reversed(invalid_rows):
        ws.delete_rows(row_num)

    # Format headers
    ws.cell(1, 1).font = openpyxl.styles.Font(bold=True)
    ws.auto_filter.ref = ws.dimensions

    print(f"Extracted {len(invalid_rows)} invalid rows to 'Invalid' sheet")
    return len(invalid_rows)

def assign_staff(ws, date_token: str = None, include_programming: bool = False,
                 assign_cathy: bool = False, cathy_all_payers: bool = False,
                 skip_rosanna: bool = False):
    """Assign staff names based on the standard daily billing rules.

    - Self Pay (GROUPFLD2 == "Self Pay") always goes to CB; every service
      bills every day, with no exceptions.
    - Jasmine and Rosanna only ever receive GROUPFLD2 == "Insurance" rows.
    - Among Insurance rows, the "professional pool" is every row whose
      Claim Type is CMS-1500 or UB-04 (UB-04 counts as Professional every
      day), sorted alphabetically by Client. Monday through Friday,
      Rosanna receives the first ROSANNA_PROFESSIONAL_CAP[weekday] (150)
      of that sorted pool; the rest of the pool goes to Jasmine. Rosanna
      caps no rows on weekends, so Jasmine gets the whole pool those days.
    - Jasmine also receives Insurance rows that are billable Programming
      (Detox, Residential) or e-care for that weekday, as well as any
      other billable Insurance row whose Claim Type is not CMS-1500/UB-04
      (i.e. institutional/837I), unless it's PHP (always Melissa's).
    - IOP (including Telemed IOP) always goes to Jasmine, every day of the
      week, bypassing the professional pool/Rosanna split even if Claim
      Type is CMS-1500 or UB-04 — unless the Cathy report is on and the
      row is a Professional row for one of her payers, which is hers.
    - Any other Insurance row (not billable that day), or any row that is
      neither Self Pay nor Insurance, is Unable to Bill.
    - Melissa (WM Program Level, PHP/Partial Hospitalization, or
      Aetna/Humana Detox/Residential) and the O'Flynn Karen OP
      Chappaqua/OP NYC "Unable to Bill" rule take priority over all of the
      above. PHP rows are assigned to Melissa every day in the Masters
      spreadsheet; she does not get an individual report, and PHP is
      billed only on Tuesdays as an operational matter.

    Four optional, per-run overrides (all off by default) sit on top of the
    schedule above:
    - include_programming: Programming (Detox, Residential) is billable
      regardless of the weekday, so it can be worked on a Monday or
      Wednesday. E-care is unaffected and stays Tuesday-only.
    - assign_cathy: every Professional (CMS-1500/UB-04) Insurance row whose
      Payer is Oxford, ConnectiCare, or UBH goes to Cathy instead of into
      the Rosanna/Jasmine professional pool, so no row is worked twice.
      Whatever the service is, it is hers: that includes IOP for those
      three payers, which she takes ahead of the IOP-to-Jasmine rule. The
      rules ahead of her (Self Pay, WM, O'Flynn Karen, Aetna/Humana
      Detox/Residential, and PHP) still take priority.
    - cathy_all_payers: run that same Cathy rule against her full payer
      list (CATHY_ALL_PAYERS) instead of just her usual three — adding
      Emblem, Surest, UBH-HP and UMR. Only the payer list widens: it is
      still Professional Insurance rows only, and the same rules still
      take priority over her. This turns the Cathy report on by itself,
      whether or not assign_cathy is also set.
    - skip_rosanna: Rosanna is given no rows at all. The whole professional
      pool left after Cathy's goes to Jasmine, exactly as it does on a
      weekend.
    """

    cols = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        if header == "GROUPFLD2":
            cols['group'] = col
        elif header == "GROUPFLD1":
            cols['group_fld1'] = col
        elif header == "Service":
            cols['service'] = col
        elif header == "Payer":
            cols['payer'] = col
        elif header == "Billing Provider":
            cols['billing_provider'] = col
        elif header == "Program Level":
            cols['program_level'] = col
        elif header == "Client":
            cols['client'] = col
        elif header == "Claim Type":
            cols['claim_type'] = col

    if not all(k in cols for k in ['group', 'service', 'payer', 'claim_type', 'client']):
        raise ValueError(
            "Missing required columns: GROUPFLD2, Service, Payer, Claim Type, or Client"
        )

    weekday, did_fallback = parse_weekday_from_token(date_token)
    if did_fallback:
        print(f"Failed to parse date_token '{date_token}', using current weekday={weekday}")
    else:
        print(f"Using date from filename: {date_token} (weekday={weekday})")

    # "All of her payers" implies the Cathy report: turning that option on
    # alone is enough to run it.
    assign_cathy = assign_cathy or cathy_all_payers
    is_cathy_row_payer = is_cathy_all_payer if cathy_all_payers else is_cathy_payer

    # Rosanna takes nothing when skip_rosanna is on, so the whole pool falls
    # to Jasmine — the same path a weekend already takes.
    rosanna_cap = 0 if skip_rosanna else ROSANNA_PROFESSIONAL_CAP.get(weekday, 0)
    if rosanna_cap:
        capped_staff, professional_cap = "Rosanna", rosanna_cap
    else:
        capped_staff, professional_cap = None, 0

    row_data_map = {}
    fixed_staff = {}
    professional_rows = []
    other_rows = []

    for row in range(2, ws.max_row + 1):
        row_data_map[row] = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]

        group = str(ws.cell(row, cols['group']).value or "").strip()
        service = str(ws.cell(row, cols['service']).value or "")
        payer = str(ws.cell(row, cols['payer']).value or "").lower()
        claim_type = str(ws.cell(row, cols['claim_type']).value or "")

        if group == "Self Pay":
            fixed_staff[row] = "CB"
            other_rows.append(row)
            continue

        staff = None

        if 'program_level' in cols and is_wm_program_level(ws.cell(row, cols['program_level']).value):
            staff = "Melissa"

        if not staff and 'billing_provider' in cols and 'group_fld1' in cols:
            billing_provider = str(ws.cell(row, cols['billing_provider']).value or "").strip()
            group_fld1 = str(ws.cell(row, cols['group_fld1']).value or "").strip()
            if billing_provider == "O'Flynn, Karen" and group_fld1 in ("OP Chappaqua", "OP NYC"):
                staff = "Unable to Bill"

        if not staff:
            service_lower = service.lower()
            has_detox_res = ("detox" in service_lower or "residential" in service_lower)
            has_insurance = "aetna" in payer or "humana" in payer
            if has_detox_res and has_insurance and not is_drug_screen(service):
                staff = "Melissa"

        if not staff and is_php_service(service):
            staff = "Melissa"

        if staff:
            fixed_staff[row] = staff
            other_rows.append(row)
            continue

        if group != "Insurance":
            fixed_staff[row] = "Unable to Bill"
            other_rows.append(row)
            continue

        # Cathy (optional): every Professional row for her payers is hers,
        # whatever the service is. That includes IOP, so this sits ahead of
        # the IOP-to-Jasmine rule below. Her rows leave the Rosanna/Jasmine
        # professional pool entirely rather than being worked twice.
        if (assign_cathy and is_professional_claim_type(claim_type)
                and is_cathy_row_payer(payer)):
            fixed_staff[row] = "Cathy"
            other_rows.append(row)
            continue

        # IOP (including Telemed IOP) always goes to Jasmine, every day of
        # the week, bypassing the professional pool/Rosanna split even if
        # Claim Type is CMS-1500 or UB-04.
        if is_iop_service(service):
            fixed_staff[row] = "Jasmine"
            other_rows.append(row)
            continue

        if is_professional_claim_type(claim_type):
            client = str(ws.cell(row, cols['client']).value or "").strip()
            professional_rows.append((row, client))
            continue

        if is_non_billable_service_for_weekday(
                service, weekday, include_programming=include_programming):
            fixed_staff[row] = "Unable to Bill"
        else:
            fixed_staff[row] = "Jasmine"
        other_rows.append(row)

    professional_rows.sort(key=lambda x: x[1].lower())

    ordered_rows = [row for row, _ in professional_rows] + other_rows
    new_row_pos = 2
    for original_row in ordered_rows:
        for col in range(1, ws.max_column + 1):
            ws.cell(new_row_pos, col).value = row_data_map[original_row][col - 1]
        new_row_pos += 1

    num_capped = min(professional_cap, len(professional_rows))
    new_row_pos = 2
    for idx in range(len(professional_rows)):
        ws.cell(new_row_pos, 1).value = capped_staff if (capped_staff and idx < num_capped) else "Jasmine"
        new_row_pos += 1
    for original_row in other_rows:
        ws.cell(new_row_pos, 1).value = fixed_staff[original_row]
        new_row_pos += 1

    print("Staff assignment complete")

def finalize_workbook(wb, include_batch_billings: bool = False,
                      include_iop_status: bool = False):
    """Add Status/Comments columns and validation for Rosanna/Jasmine/Cathy exports.

    Jasmine and Cathy share the same Status dropdown, which adds 'Batch
    Billings' and 'IOP' to the list Rosanna gets.
    """
    ws = wb.active

    # Insert two columns at E
    ws.insert_cols(5, 2)
    ws.cell(1, 5).value = "Status"
    ws.cell(1, 6).value = "Comments"

    # Create Sheet2 with validation list
    ws_list = wb.create_sheet("Sheet2")
    status_items = ["Billed", "Unable to Bill", "Contractual Adj", "Incomplete Billings",
                    "Utox Batch", "Inclusive Services"]
    if include_batch_billings:
        status_items.append("Batch Billings")
    if include_iop_status:
        status_items.append("IOP")

    for idx, item in enumerate(status_items, start=1):
        ws_list[f'A{idx}'] = item
    list_range = f"=Sheet2!$A$1:$A${len(status_items)}"

    # Add data validation to Status column
    dv = DataValidation(type="list", formula1=list_range, allow_blank=True)
    ws.add_data_validation(dv)

    last_row = ws.max_row
    dv.add(f'E2:E{last_row}')

    # Auto-fit columns
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

def export_staff_workbooks(wb, wb_path, date_token, exclude_aetna: bool = False,
                           cathy_report: bool = False, skip_rosanna: bool = False):
    """Export separate workbooks for Rosanna, Jasmine, and CB.

    All staff (Melissa, Unable to Bill, etc.) are still assigned in the
    Masters workbook, but only Rosanna, Jasmine, and CB receive individual
    reports.

    Args:
        exclude_aetna: When True, Aetna rows are left out of every individual
            workbook. They still appear in the Masters workbook.
        cathy_report: When True, also export Cathy's workbook (the
            Professional-only rows for her payers that assign_staff routed
            to her).
        skip_rosanna: When True, no workbook is exported for Rosanna — she
            was assigned nothing for this run.
    """

    ws = wb.active
    save_folder = get_save_folder(wb_path, date_token)

    # Find Program Level column for WM filtering, and Payer for the Aetna
    # exclusion.
    program_level_col = None
    payer_col = None
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        if header == "Program Level":
            program_level_col = col
        elif header == "Payer":
            payer_col = col

    # Count and warn about WM rows
    wm_count = 0
    if program_level_col is not None:
        for row in range(2, ws.max_row + 1):
            if is_wm_program_level(ws.cell(row, program_level_col).value):
                wm_count += 1
    if wm_count > 0:
        print(
            f"WARNING: {wm_count} row(s) with 'WM' in the Program Level column were found. "
            "These rows are assigned to Melissa and appear only in the Masters report. "
            "Only Melissa is authorized to bill WM."
        )

    staff_reports = ["Jasmine", "CB"] if skip_rosanna else ["Rosanna", "Jasmine", "CB"]
    if cathy_report:
        staff_reports.append("Cathy")

    for staff_name in staff_reports:
        # Create new workbook
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = "Sheet1"

        # Copy header
        for col in range(1, ws.max_column + 1):
            new_ws.cell(1, col).value = ws.cell(1, col).value

        # Copy matching rows
        new_row = 2
        for row in range(2, ws.max_row + 1):
            assigned_staff = ws.cell(row, 1).value
            if assigned_staff == staff_name:
                # Exclude Aetna rows from every individual workbook; they are
                # still assigned in the Masters workbook.
                if exclude_aetna and payer_col is not None:
                    if is_aetna_payer(str(ws.cell(row, payer_col).value or "")):
                        continue
                # Skip WM program level rows for all staff except Melissa
                # (Masters retains all rows; only Melissa bills WM)
                if (staff_name != "Melissa" and
                        program_level_col is not None and
                        is_wm_program_level(ws.cell(row, program_level_col).value)):
                    continue
                for col in range(1, ws.max_column + 1):
                    new_ws.cell(new_row, col).value = ws.cell(row, col).value
                new_row += 1

        if new_row == 2:  # No rows found
            print(f"No rows for {staff_name}, skipping")
            continue

        # Finalize for Rosanna, Jasmine, and Cathy only. Cathy's Status
        # dropdown carries the same options as Jasmine's.
        if staff_name in ["Rosanna", "Jasmine", "Cathy"]:
            jasmine_options = staff_name in ("Jasmine", "Cathy")
            finalize_workbook(new_wb, include_batch_billings=jasmine_options,
                              include_iop_status=jasmine_options)

        # Save
        save_path = save_folder / f"Unbilled Revenue By Resident and Funding Type {date_token} - {staff_name}.xlsx"
        new_wb.save(save_path)
        print(f"Saved {save_path}")

def main(workbook_path, include_programming: bool = False, exclude_aetna: bool = False,
         cathy_report: bool = False, cathy_all_payers: bool = False,
         skip_rosanna: bool = False):
    """Main workflow.

    The five optional flags mirror the checkboxes in the Streamlit app and
    are all off by default, so a plain run follows the standard daily
    schedule:
      include_programming - bill Programming (Detox/Residential) regardless
          of the weekday.
      exclude_aetna       - keep Aetna rows out of the individual workbooks.
      cathy_report        - route Professional Oxford/ConnectiCare/UBH rows
          to Cathy and save her workbook.
      cathy_all_payers    - run the Cathy report against her full payer list
          (CATHY_ALL_PAYERS) instead of her usual three; turns the report on
          by itself.
      skip_rosanna        - give Rosanna nothing: Jasmine takes the whole
          professional pool and no workbook is saved for Rosanna.
    """
    # "All of her payers" runs the Cathy report on its own.
    cathy_report = cathy_report or cathy_all_payers

    # Load workbook
    wb = openpyxl.load_workbook(workbook_path)

    # Find the data sheet (first sheet or by name pattern)
    ws = wb.active

    # Extract date token
    date_token = extract_date_from_filename(Path(workbook_path).name)
    if not date_token:
        print("Warning: Could not extract date (MMDDYYYY) from filename, will use today's date")

    print(f"Processing file with date: {date_token if date_token else 'today'}")

    # Step 1: Extract invalid
    step_1_extract_invalid(ws)

    # Step 2-6: Assign staff
    assign_staff(ws, date_token, include_programming=include_programming,
                 assign_cathy=cathy_report, cathy_all_payers=cathy_all_payers,
                 skip_rosanna=skip_rosanna)

    # Step 7: Export individual workbooks (only if date_token is available)
    if date_token:
        export_staff_workbooks(wb, workbook_path, date_token,
                               exclude_aetna=exclude_aetna,
                               cathy_report=cathy_report,
                               skip_rosanna=skip_rosanna)
    else:
        print("Skipping individual workbook export due to missing date token")

    # Save main workbook
    wb.save(workbook_path)
    print(f"Saved main workbook: {workbook_path}")

    print("\n✓ All done!")

if __name__ == "__main__":
    # Optionally allow running from CLI
    import argparse

    parser = argparse.ArgumentParser(
        description="Process an unbilled revenue workbook."
    )
    parser.add_argument("workbook_path", nargs="?",
                        help="Path to the .xlsx workbook to process.")
    parser.add_argument("--include-programming", action="store_true",
                        help="Bill Programming (Detox/Residential) regardless of "
                             "the weekday, so it can be worked on a Monday or "
                             "Wednesday.")
    parser.add_argument("--exclude-aetna", action="store_true",
                        help="Keep Aetna rows out of the individual staff "
                             "workbooks (they stay in the Masters workbook).")
    parser.add_argument("--cathy-report", action="store_true",
                        help="Assign Professional (CMS-1500/UB-04) Insurance rows "
                             "for " + ", ".join(CATHY_PAYERS) + " to Cathy and "
                             "save her workbook.")
    parser.add_argument("--cathy-all-payers", action="store_true",
                        help="Run the Cathy report against her full payer list ("
                             + ", ".join(CATHY_ALL_PAYERS) + ") instead of just "
                             "her usual three. Turns the Cathy report on by "
                             "itself; --cathy-report is not also needed.")
    parser.add_argument("--no-rosanna", action="store_true", dest="skip_rosanna",
                        help="Give Rosanna nothing for this run: Jasmine takes the "
                             "whole Professional pool and no workbook is saved for "
                             "Rosanna.")
    args = parser.parse_args()

    if args.workbook_path:
        main(args.workbook_path,
             include_programming=args.include_programming,
             exclude_aetna=args.exclude_aetna,
             cathy_report=args.cathy_report,
             cathy_all_payers=args.cathy_all_payers,
             skip_rosanna=args.skip_rosanna)
