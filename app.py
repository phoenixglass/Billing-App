import streamlit as st
import openpyxl
import re
from datetime import datetime
import io
import tempfile
import os
import json
import logging
from pathlib import Path

from billing_rules import (
    is_aetna_payer,
    is_wm_program_level,
    is_anthem_payer,
    is_bcb_anthem_ct_php_res_detox,
    _is_drug_screen as is_drug_screen,
    CATHY_PAYERS,
    CATHY_ALL_PAYERS,
    validate_custom_report_name,
    parse_terms,
    matches_any_term,
    assign_staff,
    finalize_workbook,
)

# GCS logging is optional — the app falls back to local logging without it, so
# a problem importing it must never stop the app from starting.
#
# ImportError alone is not enough. A broken (rather than missing) install fails
# in other ways: a partially installed wheel raises OSError/AttributeError, and
# a mismatched `cryptography` build raises pyo3's PanicException, which derives
# from BaseException and so slips past `except Exception`. Catch BaseException
# and re-raise only genuine control-flow exceptions.
try:
    from google.cloud import storage
    from google.oauth2 import service_account
    GCS_AVAILABLE = True
except (KeyboardInterrupt, SystemExit):
    raise
except BaseException as e:  # noqa: BLE001 - optional dependency must not break boot
    print(f"WARNING: Google Cloud Storage libraries unavailable "
          f"({type(e).__name__}: {e}). Using local audit logging only.")
    GCS_AVAILABLE = False

# Configure audit logging
LOG_DIR = Path("audit_logs")


def _ensure_log_dir() -> bool:
    """Create the audit log directory, returning False if it is unusable.

    Audit logging must never prevent the app from starting: on a read-only or
    full filesystem we fall back to console logging, which the hosting platform
    still captures.
    """
    try:
        LOG_DIR.mkdir(exist_ok=True)
        return True
    except OSError as e:
        print(f"WARNING: Cannot create audit log directory '{LOG_DIR}': {e}. "
              "Falling back to console logging only.")
        return False

def setup_logging():
    """Setup logging with GCS support if credentials are available."""
    handlers = []
    gcs_handler = None

    # Try to setup GCS logging (for Streamlit Cloud)
    if GCS_AVAILABLE and "gcp_service_account" in st.secrets:
        try:
            creds_json = st.secrets["gcp_service_account"]
            if isinstance(creds_json, str):
                creds_dict = json.loads(creds_json)
            else:
                creds_dict = dict(creds_json)

            credentials = service_account.Credentials.from_service_account_info(creds_dict)
            gcs_client = storage.Client(credentials=credentials)
            bucket_name = st.secrets.get("gcs_bucket_name", "billing-app-audit-logs")
            bucket = gcs_client.bucket(bucket_name)

            # Verify bucket exists
            if bucket.exists():
                gcs_handler = GCSHandler(gcs_client, bucket_name)
                handlers.append(gcs_handler)
            else:
                print(f"WARNING: GCS bucket '{bucket_name}' not found. Using local logging only.")
        except Exception as e:
            print(f"WARNING: Failed to setup GCS logging: {e}. Using local logging only.")

    # Local file handler for development/fallback. A failure to open the log
    # file must not stop startup — console logging below still reaches the
    # hosting platform's log capture.
    if _ensure_log_dir():
        try:
            log_file = LOG_DIR / f"audit_{datetime.now().strftime('%Y%m%d')}.log"
            handlers.append(logging.FileHandler(log_file))
        except OSError as e:
            print(f"WARNING: Cannot open audit log file for writing: {e}. "
                  "Falling back to console logging only.")

    # Add console handler
    handlers.append(logging.StreamHandler())

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=handlers,
        force=True
    )

    # Security: Enforce restrictive permissions on audit logs (owner read/write
    # only). Best-effort: some filesystems reject chmod, which must not stop
    # startup. The failure is logged so it is visible during a review.
    try:
        for existing_log in LOG_DIR.glob("audit_*.log"):
            os.chmod(existing_log, 0o600)
    except OSError as e:
        print(f"WARNING: Could not restrict audit log permissions: {e}")

    return gcs_handler

class GCSHandler(logging.Handler):
    """Custom logging handler that writes to Google Cloud Storage."""

    def __init__(self, gcs_client, bucket_name):
        super().__init__()
        self.gcs_client = gcs_client
        self.bucket_name = bucket_name
        self.bucket = gcs_client.bucket(bucket_name)

    def emit(self, record):
        """Write log record to GCS."""
        try:
            log_entry = self.format(record) + "\n"
            log_date = datetime.now().strftime('%Y%m%d')
            blob_name = f"audit_logs/audit_{log_date}.log"

            blob = self.bucket.blob(blob_name)

            # Append to existing log
            if blob.exists():
                existing = blob.download_as_string().decode('utf-8')
                blob.upload_from_string(existing + log_entry)
            else:
                blob.upload_from_string(log_entry)
        except Exception as e:
            self.handleError(record)

gcs_handler = setup_logging()
logger = logging.getLogger(__name__)

if gcs_handler:
    logger.info("✓ GCS logging enabled")

st.set_page_config(page_title="Unbilled Billing App", layout="wide")

st.title("Unbilled Billing Processor")
st.markdown("Upload your billing file to process it automatically")


def extract_date_from_filename(filename):
    """Extract MMDDYYYY date from filename"""
    match = re.search(r'(\d{8})', filename)
    if match:
        return match.group(1)
    return None

def get_filename_prefix(filename):
    """Extract the prefix before the date in filename"""
    date_match = re.search(r'(\d{8})', filename)
    if date_match:
        date_pos = date_match.start()
        date_str = date_match.group(1)
        prefix = filename[:date_pos + len(date_str)]
        remaining = filename[date_pos + len(date_str):]
        if remaining.startswith(' - '):
            prefix += ' - '
        return prefix
    return ""


def step_1_extract_invalid(ws):
    """Insert Staff/Status, extract Invalid rows to new sheet, delete from main"""

    if ws.cell(1, 1).value != "Staff/Status":
        ws.insert_cols(1)
        ws.cell(1, 1).value = "Staff/Status"

    billable_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(1, col).value == "Billable Status":
            billable_col = col
            break

    if not billable_col:
        raise ValueError("Billable Status column not found")

    invalid_rows = []
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, billable_col).value == "Invalid for Billing":
            invalid_rows.append(row)

    if invalid_rows:
        wb = ws.parent
        if "Invalid" in wb.sheetnames:
            del wb["Invalid"]

        ws_invalid = wb.create_sheet("Invalid")

        for col in range(1, ws.max_column + 1):
            ws_invalid.cell(1, col).value = ws.cell(1, col).value

        for idx, row_num in enumerate(invalid_rows, start=2):
            for col in range(1, ws.max_column + 1):
                ws_invalid.cell(idx, col).value = ws.cell(row_num, col).value

        for row_num in reversed(invalid_rows):
            ws.delete_rows(row_num)

    for col in range(1, ws.max_column + 1):
        ws.cell(1, col).font = openpyxl.styles.Font(bold=True)
    ws.auto_filter.ref = ws.dimensions

    return len(invalid_rows) if invalid_rows else 0

def validate_uploaded_file(uploaded_file):
    """Validate uploaded file for security."""
    # Check file size (max 50MB)
    MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB
    file_size = len(uploaded_file.getvalue())
    
    if file_size > MAX_FILE_SIZE:
        raise ValueError(f"File too large ({file_size/1024/1024:.1f}MB). Maximum allowed: 50MB")
    
    # Validate file extension
    if not uploaded_file.name.endswith('.xlsx'):
        raise ValueError("Only .xlsx files are allowed")
    
    # Check for suspicious patterns in filename
    suspicious_patterns = ['..', '/', '\\', '<', '>', '|', ':', '*', '?', '"']
    if any(pattern in uploaded_file.name for pattern in suspicious_patterns):
        raise ValueError("Filename contains invalid characters")
    
    logger.info(f"Uploaded file: {uploaded_file.name} ({file_size} bytes)")
    return True

def is_op_wm_program_level(cell_value) -> bool:
    """Return True if the Program Level cell contains 'OP WM'."""
    return "OP WM" in str(cell_value or "").upper()

def secure_cleanup(file_path):
    """Securely delete temporary file."""
    try:
        if os.path.exists(file_path):
            # Overwrite with random data before deletion (simple secure delete)
            with open(file_path, 'ba+', buffering=0) as f:
                length = f.tell()
                f.seek(0)
                f.write(os.urandom(length))
            os.remove(file_path)
            logger.info(f"Securely deleted temp file: {file_path}")
    except Exception as e:
        logger.error(f"Error during secure cleanup: {e}")

def process_workbook(uploaded_file, exclude_optum: bool = False,
                     exclude_bcb_anthem_ct: bool = False,
                     exclude_anthem_rosanna_jasmine_owm: bool = False,
                     exclude_detox_residential: bool = False,
                     include_programming: bool = False,
                     exclude_aetna: bool = False,
                     cathy_report: bool = False,
                     cathy_all_payers: bool = False,
                     skip_rosanna: bool = False,
                     exclude_payer_terms: list = None,
                     exclude_service_terms: list = None,
                     exclude_scope: list = None,
                     rosanna_cap_override: int = None,
                     custom_report_name: str = None,
                     custom_report_payer_terms: list = None,
                     custom_report_professional_only: bool = True):
    """Process the uploaded workbook.

    The four exclude_* flags and include_programming/cathy_report/
    cathy_all_payers/skip_rosanna are all per-run options driven by the
    checkboxes below; every one of them is off by default, so an unchecked
    run follows the standard daily schedule.

    exclude_payer_terms/exclude_service_terms are the free-text custom
    exclusion fields: any row whose Payer or Service contains one of these
    terms (case-insensitive) is left out of the individual workbooks for
    this run, without needing a new checkbox or a code change. exclude_scope
    limits which staff's workbooks the two custom exclusions apply to; an
    empty/None scope applies them to every individual workbook, matching
    how exclude_aetna and the other blanket exclusions behave.

    rosanna_cap_override, when set, replaces the standard weekday cap with
    this exact row count for the run (ignored if skip_rosanna is set).

    custom_report_name/custom_report_payer_terms/
    custom_report_professional_only configure a second, generic Cathy-shaped
    report: matching rows go to a new named staff member with their own
    workbook, without a code change. See assign_staff's docstring for the
    exact placement/priority rules.
    """
    # "Cathy report: all of her payers" runs her report by itself, so the
    # operator only has to check the one box.
    cathy_report = cathy_report or cathy_all_payers
    if custom_report_name:
        validate_custom_report_name(custom_report_name)
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(uploaded_file.getbuffer())
            tmp_path = tmp.name

        logger.info(f"Processing file: {uploaded_file.name} (temp: {tmp_path})")

        wb = openpyxl.load_workbook(tmp_path)
        ws = wb.active

        date_token = extract_date_from_filename(uploaded_file.name)
        if not date_token:
            raise ValueError("Filename must contain 8 digits (MMDDYYYY)")

        filename_prefix = get_filename_prefix(uploaded_file.name)

        invalid_count = step_1_extract_invalid(ws)
        assign_staff(ws, date_token, include_programming=include_programming,
                     assign_cathy=cathy_report,
                     cathy_all_payers=cathy_all_payers,
                     skip_rosanna=skip_rosanna,
                     rosanna_cap_override=rosanna_cap_override,
                     custom_report_name=custom_report_name,
                     custom_report_payer_terms=custom_report_payer_terms,
                     custom_report_professional_only=custom_report_professional_only)

        output_files = {}

        # Find the Service, Program Level, and Payer column indices
        service_col = None
        program_level_col = None
        payer_col = None
        for col in range(1, ws.max_column + 1):
            header = ws.cell(1, col).value
            if header == "Service":
                service_col = col
            elif header == "Program Level":
                program_level_col = col
            elif header == "Payer":
                payer_col = col

        # Count WM rows for alerts (both WM and OP WM go to Melissa)
        wm_count = 0
        if program_level_col is not None:
            for row in range(2, ws.max_row + 1):
                pl = ws.cell(row, program_level_col).value
                if is_wm_program_level(pl):
                    wm_count += 1

        # Rosanna, Jasmine, and CB always get individual reports (empty ones
        # are skipped below); Cathy gets one only when her report is turned
        # on for this run, the custom report's staff gets one only when it's
        # configured, and Rosanna none at all when she is skipped. Melissa,
        # Unable to Bill, etc. stay Masters-only.
        staff_reports = ["Jasmine", "CB"] if skip_rosanna else ["Rosanna", "Jasmine", "CB"]
        if cathy_report:
            staff_reports.append("Cathy")
        if custom_report_name:
            staff_reports.append(custom_report_name)

        for staff_name in staff_reports:
            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            new_ws.title = "Sheet1"

            for col in range(1, ws.max_column + 1):
                new_ws.cell(1, col).value = ws.cell(1, col).value

            new_row = 2
            for row in range(2, ws.max_row + 1):
                assigned_staff = ws.cell(row, 1).value
                if assigned_staff == staff_name:
                    # Exclude Optum utox rows from all individual workbooks
                    if (exclude_optum and service_col is not None and
                            payer_col is not None):
                        service_val = str(ws.cell(row, service_col).value or "")
                        payer_val = str(ws.cell(row, payer_col).value or "").lower()
                        if is_drug_screen(service_val) and "optum" in payer_val:
                            continue
                    if (exclude_bcb_anthem_ct and service_col is not None and
                            payer_col is not None):
                        service_val = str(ws.cell(row, service_col).value or "")
                        payer_val = str(ws.cell(row, payer_col).value or "")
                        if is_bcb_anthem_ct_php_res_detox(payer_val, service_val):
                            continue
                    if (exclude_anthem_rosanna_jasmine_owm and
                            assigned_staff in ("Rosanna", "Jasmine") and
                            payer_col is not None):
                        payer_val = str(ws.cell(row, payer_col).value or "")
                        if is_anthem_payer(payer_val):
                            continue
                    # Exclude Aetna rows from every individual workbook;
                    # they are still assigned in the Masters report.
                    if exclude_aetna and payer_col is not None:
                        payer_val = str(ws.cell(row, payer_col).value or "")
                        if is_aetna_payer(payer_val):
                            continue
                    if exclude_detox_residential and service_col is not None:
                        service_val = str(ws.cell(row, service_col).value or "").lower()
                        if "detox" in service_val or "residential" in service_val:
                            continue
                    # Custom per-run exclusions (free text, no code change
                    # needed): skip if this staff is in scope (or scope is
                    # empty, meaning everyone) and the payer/service matches.
                    in_scope = not exclude_scope or assigned_staff in exclude_scope
                    if in_scope and exclude_payer_terms and payer_col is not None:
                        payer_val = str(ws.cell(row, payer_col).value or "")
                        if matches_any_term(payer_val, exclude_payer_terms):
                            continue
                    if in_scope and exclude_service_terms and service_col is not None:
                        service_val = str(ws.cell(row, service_col).value or "")
                        if matches_any_term(service_val, exclude_service_terms):
                            continue
                    # Skip WM/OP WM program level rows for all staff except Melissa
                    if (staff_name != "Melissa" and
                            program_level_col is not None and
                            is_wm_program_level(ws.cell(row, program_level_col).value)):
                        continue
                    for col in range(1, ws.max_column + 1):
                        new_ws.cell(new_row, col).value = ws.cell(row, col).value
                    new_row += 1

            if new_row == 2:
                continue

            if staff_name == "CB":
                finalize_workbook(new_wb, skip_status_columns=True)
            else:
                # Rosanna gets the plain dropdown; Jasmine, Cathy, and a
                # custom report's staff all share the wider one with Batch
                # Billings and IOP included.
                jasmine_options = staff_name != "Rosanna"
                finalize_workbook(new_wb, include_batch_billings=jasmine_options,
                                  include_iop_status=jasmine_options)

            output = io.BytesIO()
            new_wb.save(output)
            output.seek(0)
            output_filename = f"{filename_prefix}{staff_name}.xlsx"
            output_files[output_filename] = output

        main_output = io.BytesIO()
        wb.save(main_output)
        main_output.seek(0)
        main_filename = f"{filename_prefix}Masters.xlsx"
        output_files[main_filename] = main_output

        logger.info(f"Successfully processed file: {uploaded_file.name}, generated {len(output_files)} output files")

        return output_files, invalid_count, date_token, wm_count

    finally:
        # Always cleanup temp file securely
        if tmp_path:
            secure_cleanup(tmp_path)

# UI
st.markdown("### Upload your billing file")
st.info("🔒 All actions are logged. Session timeout: 15 minutes of inactivity.")

uploaded_file = st.file_uploader(
    "Choose an Excel file (must have MMDDYYYY in filename)",
    type="xlsx",
    key="billing_file_uploader"
)

# Preserve uploaded file in session state to prevent it from disappearing on reruns
if uploaded_file is not None:
    st.session_state["current_uploaded_file"] = uploaded_file

st.markdown(
    "Staff assignment now follows the standard daily schedule automatically, based on "
    "the date in the filename: Self Pay always bills every day to **CB**; for Insurance "
    "rows, **Rosanna** gets the first 150 Professional (Claim Type CMS-1500 or UB-04) "
    "services each weekday (Monday-Friday), sorted alphabetically by Client, and "
    "**Jasmine** gets the remaining Professional rows plus any billable Programming/e-care "
    "rows for that day. PHP rows always go to **Melissa** in the Masters report (no "
    "individual report is generated for her)."
)

exclude_optum = st.checkbox(
    "Exclude Optum insurance",
    value=False,
    help="When checked, utox (drug screen) rows with Optum as the payer will be excluded from all individual staff workbooks."
)

exclude_bcb_anthem_ct = st.checkbox(
    "Exclude BCB Anthem CT for PHP, Residential, and Detox",
    value=False,
    help="When checked, rows where the payer is BCB Anthem CT and the service is PHP (Partial Hospitalization), Residential, or Detox will be excluded from all individual staff workbooks."
)

exclude_anthem_rosanna_jasmine_owm = st.checkbox(
    "Remove Anthem from Rosanna and Jasmine reports",
    value=False,
    help="When checked, rows where the payer contains 'Anthem' will be excluded from Rosanna's and Jasmine's workbooks. Anthem rows are still retained in the Masters report."
)

exclude_detox_residential = st.checkbox(
    "Don't give anyone Detox or Residential",
    value=False,
    help="When checked, Detox and Residential service rows are excluded from all individual staff workbooks. They still appear in the Masters report."
)

st.markdown("**One-off options for this run**")

include_programming = st.checkbox(
    "Include Programming (Detox/Residential) today",
    value=False,
    help=(
        "When checked, Programming (Detox, Residential) is billable no matter what "
        "day it is, so it can be worked on a Monday or Wednesday. Billable "
        "Programming rows go to Jasmine as usual. E-care is unaffected and stays "
        "Tuesday-only."
    )
)

exclude_aetna = st.checkbox(
    "Exclude Aetna",
    value=False,
    help=(
        "When checked, every row whose payer is Aetna is excluded from all "
        "individual staff workbooks. Aetna rows still appear in the Masters report."
    )
)

cathy_report = st.checkbox(
    "Cathy report: Professional services only for "
    + ", ".join(CATHY_PAYERS),
    value=False,
    help=(
        "When checked, Insurance rows whose Claim Type is Professional (CMS-1500 or "
        "UB-04) and whose payer is Oxford, ConnectiCare, or UBH are assigned to "
        "Cathy and saved as her own workbook, whatever the service is — IOP for "
        "those payers is hers too. Those rows leave the Rosanna/Jasmine "
        "professional pool, so no row is worked twice — Rosanna's 150-row cap then "
        "applies to what is left. WM, PHP, and the O'Flynn Karen rule still take "
        "priority over Cathy."
    )
)

cathy_all_payers = st.checkbox(
    "Cathy report: all of her payers ("
    + ", ".join(CATHY_ALL_PAYERS)
    + ")",
    value=False,
    help=(
        "The same Cathy report, run against her full payer list instead of just "
        "her usual three: it adds Emblem, Surest, UBH-HP, and UMR. Only the payer "
        "list widens — it is still Professional (CMS-1500/UB-04) Insurance rows "
        "only, they still leave the Rosanna/Jasmine pool so no row is worked "
        "twice, and WM, PHP, and the O'Flynn Karen rule still take priority. "
        "Checking this runs the Cathy report on its own; the box above does not "
        "also need to be checked."
    )
)

skip_rosanna = st.checkbox(
    "Don't give Rosanna anything",
    value=False,
    help=(
        "When checked, Rosanna is assigned no rows for this run and no workbook "
        "is generated for her. Her share of the Professional pool goes to Jasmine "
        "instead, the same way it does on a weekend. Nothing is left unassigned: "
        "every row still appears in the Masters report with an owner."
    )
)

st.markdown("**Rosanna's cap for this run (optional, no code change needed)**")

override_rosanna_cap = st.checkbox(
    "Override Rosanna's cap for today",
    value=False,
    help=(
        "The standard schedule gives Rosanna the first 150 professional-pool rows "
        "Monday-Friday and none on weekends. Check this to give her a different "
        "row count for this run only. Ignored if 'Don't give Rosanna anything' is checked."
    ),
)
rosanna_cap_override = None
if override_rosanna_cap:
    rosanna_cap_override = int(st.number_input(
        "Rosanna's row cap for this run",
        min_value=0,
        value=150,
        step=10,
        help="Rosanna receives up to this many rows from the professional pool for this run.",
    ))

st.markdown("**Custom report (optional, no code change needed) — route a payer to a new staff member**")

custom_report_name = st.text_input(
    "Staff name for this report (leave blank to skip)",
    value="",
    help=(
        "When set (with the payer list below), every Insurance row whose Payer "
        "matches goes to this staff member instead of Rosanna/Jasmine, and they "
        "get their own workbook for this run — the same way Cathy's report works, "
        "for a different payer/staff combination. Must not be Rosanna, Jasmine, "
        "CB, Melissa, Cathy, or Unable to Bill."
    ),
)

custom_report_payers_raw = st.text_input(
    "Payers for this report (comma-separated)",
    value="",
    help="Case-insensitive substring match against the Payer column. Example: Cigna, Humana",
)

custom_report_professional_only = st.checkbox(
    "Professional claim types only (CMS-1500/UB-04)",
    value=True,
    help=(
        "When checked (default, same restriction Cathy has), only Professional "
        "rows for these payers go to this report. Uncheck to match any claim type."
    ),
)

custom_report_payer_terms = parse_terms(custom_report_payers_raw)
if custom_report_name and not custom_report_payer_terms:
    st.warning("⚠️ A custom report name is set but no payers were entered, so it won't claim any rows.")
if custom_report_payer_terms and not custom_report_name:
    st.warning("⚠️ Custom report payers were entered but no staff name was given, so they'll be ignored.")

st.markdown("**Custom exclusions for this run (type a payer or service — no code change needed)**")

custom_exclude_payers_raw = st.text_input(
    "Exclude payers containing (comma-separated, optional)",
    value="",
    help=(
        "Rows whose Payer contains any of these terms (case-insensitive) are "
        "left out of the individual workbooks for this run only. They still "
        "appear in the Masters report. Example: Cigna, Humana"
    ),
)

custom_exclude_services_raw = st.text_input(
    "Exclude services containing (comma-separated, optional)",
    value="",
    help=(
        "Rows whose Service contains any of these terms (case-insensitive) are "
        "left out of the individual workbooks for this run only. They still "
        "appear in the Masters report. Example: Group Therapy"
    ),
)

custom_exclude_scope = st.multiselect(
    "Apply the two custom exclusions above only to these staff (optional)",
    options=["Rosanna", "Jasmine", "Cathy", "CB"],
    default=[],
    help="Leave empty to apply them to every individual workbook, same as Exclude Aetna above.",
)

custom_exclude_payer_terms = parse_terms(custom_exclude_payers_raw)
custom_exclude_service_terms = parse_terms(custom_exclude_services_raw)

if include_programming and exclude_detox_residential:
    st.warning(
        "⚠️ 'Include Programming (Detox/Residential) today' and \"Don't give anyone "
        "Detox or Residential\" are both checked. The exclusion wins: Programming rows "
        "will be assigned in the Masters report but kept out of every individual "
        "workbook."
    )

if uploaded_file is not None:
    try:
        validate_uploaded_file(uploaded_file)
        st.info("Processing your file...")
        output_files, invalid_count, date_token, wm_count = process_workbook(
            uploaded_file,
            exclude_optum=exclude_optum,
            exclude_bcb_anthem_ct=exclude_bcb_anthem_ct,
            exclude_anthem_rosanna_jasmine_owm=exclude_anthem_rosanna_jasmine_owm,
            exclude_detox_residential=exclude_detox_residential,
            include_programming=include_programming,
            exclude_aetna=exclude_aetna,
            cathy_report=cathy_report,
            cathy_all_payers=cathy_all_payers,
            skip_rosanna=skip_rosanna,
            rosanna_cap_override=rosanna_cap_override,
            custom_report_name=custom_report_name.strip() if custom_report_name else None,
            custom_report_payer_terms=custom_report_payer_terms,
            custom_report_professional_only=custom_report_professional_only,
            exclude_payer_terms=custom_exclude_payer_terms,
            exclude_service_terms=custom_exclude_service_terms,
            exclude_scope=custom_exclude_scope,
        )

        if wm_count > 0:
            st.warning(
                f"⚠️ Alert: {wm_count} row(s) with 'WM' or 'OP WM' in the Program Level column were found. "
                "These rows are assigned to Melissa and appear only in the Masters report. "
                "Only Melissa is authorized to bill WM."
            )

        st.success("✓ Processing complete!")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Date Extracted", date_token if date_token else "Today's date")
        with col2:
            st.metric("Invalid Rows", invalid_count or 0)
        with col3:
            st.metric("Files Generated", len(output_files))
        
        st.markdown("---")
        st.markdown("### Download Results")
        
        for output_filename, file_bytes in output_files.items():
            if st.download_button(
                label=f"Download {output_filename}",
                data=file_bytes,
                file_name=output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=output_filename
            ):
                logger.info(f"Downloaded: {output_filename}")

    except Exception as e:
        logger.error(f"Error processing file: {str(e)}")
        st.error(f"Error: {str(e)}")
