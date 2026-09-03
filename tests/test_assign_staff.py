"""
End-to-end tests for assign_staff: the Cathy report (both payer lists),
the custom report (a second, generic Cathy-shaped slot), the Rosanna cap
override, "give Rosanna nothing", and the include_programming override,
exercised against a real worksheet.

assign_staff lives in billing_rules.py; app.py and "Unbilled Step 1.py"
both import it from there. app.py also imports Streamlit, so these tests
load the standalone script instead, which re-exports the same function.

Run: python tests/test_assign_staff.py
Requires: openpyxl
"""
import importlib.util
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(REPO_ROOT))

from openpyxl import Workbook


def _load_unbilled_module():
    """Import "Unbilled Step 1.py" (its filename is not a valid module name)."""
    spec = importlib.util.spec_from_file_location(
        "unbilled_step_1", REPO_ROOT / "Unbilled Step 1.py"
    )
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


unbilled = _load_unbilled_module()
assign_staff = unbilled.assign_staff

HEADERS = ["Staff/Status", "GROUPFLD1", "GROUPFLD2", "Service", "Payer",
           "Billing Provider", "Program Level", "Client", "Claim Type"]

# MMDDYYYY tokens for the two days that matter to these tests.
WEDNESDAY = "09022026"
TUESDAY = "09012026"


def _sheet(rows):
    """Build a worksheet from (group, service, payer, client, claim_type) rows."""
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    for group, service, payer, client, claim_type in rows:
        ws.append(["", "OP Westchester", group, service, payer, "Smith, John",
                   "OP", client, claim_type])
    return ws


def _staff_by_client(ws):
    """Map Client -> assigned staff after assign_staff has reordered the rows.

    Column 1 is the Staff/Status column, so this reads the same assignment
    the Masters workbook carries for each row.
    """
    client_col = HEADERS.index("Client") + 1
    return {
        ws.cell(row, client_col).value: ws.cell(row, 1).value
        for row in range(2, ws.max_row + 1)
    }


def _cathy_candidate_rows():
    return [
        ("Insurance", "Individual Therapy", "Oxford", "Adams, Ann", "CMS-1500"),
        ("Insurance", "Individual Therapy", "ConnectiCare", "Baker, Bob", "UB-04"),
        ("Insurance", "Group Therapy", "UBH", "Carter, Cal", "CMS-1500"),
        # Not Cathy's: right payer, but not a Professional claim type.
        ("Insurance", "Individual Therapy", "Oxford", "Diaz, Dee", "837I"),
        # Not Cathy's: Professional, but a payer she does not cover.
        ("Insurance", "Individual Therapy", "Aetna", "Evans, Eve", "CMS-1500"),
        ("Insurance", "Individual Therapy", "Optum", "Frank, Fay", "CMS-1500"),
    ]


def test_cathy_off_by_default():
    """Without the flag, nobody is assigned to Cathy and the pool is unchanged."""
    ws = _sheet(_cathy_candidate_rows())
    assign_staff(ws, WEDNESDAY)
    staff = _staff_by_client(ws)

    assert "Cathy" not in staff.values()
    # Every Professional row falls to Rosanna under her weekday cap.
    for client in ("Adams, Ann", "Baker, Bob", "Carter, Cal", "Evans, Eve",
                   "Frank, Fay"):
        assert staff[client] == "Rosanna", (client, staff[client])


def test_cathy_takes_only_professional_rows_for_her_payers():
    """Oxford/ConnectiCare/UBH Professional rows go to Cathy; nothing else does."""
    ws = _sheet(_cathy_candidate_rows())
    assign_staff(ws, WEDNESDAY, assign_cathy=True)
    staff = _staff_by_client(ws)

    assert staff["Adams, Ann"] == "Cathy"
    assert staff["Baker, Bob"] == "Cathy"
    assert staff["Carter, Cal"] == "Cathy"

    # 837I Oxford is not Professional, so it is not Cathy's; on a Wednesday
    # it is not billable either.
    assert staff["Diaz, Dee"] == "Unable to Bill"
    # Professional rows for other payers stay in the Rosanna/Jasmine pool.
    assert staff["Evans, Eve"] == "Rosanna"
    assert staff["Frank, Fay"] == "Rosanna"


def test_cathy_rows_leave_the_professional_pool():
    """A row assigned to Cathy is not also given to Rosanna or Jasmine."""
    ws = _sheet(_cathy_candidate_rows())
    assign_staff(ws, WEDNESDAY, assign_cathy=True)

    assignments = [ws.cell(row, 1).value for row in range(2, ws.max_row + 1)]
    assert assignments.count("Cathy") == 3
    # Six rows in, six rows out, each with exactly one owner.
    assert len(assignments) == 6
    assert all(value for value in assignments)


def test_higher_priority_rules_still_win_over_cathy():
    """Self Pay, WM, and PHP keep their owners even for Cathy's payers."""
    ws = _sheet([
        ("Self Pay", "Individual Therapy", "Oxford", "Self, Sam", "CMS-1500"),
        ("Insurance", "Individual Therapy", "Oxford", "Wm, Wes", "CMS-1500"),
        ("Insurance", "Partial Hospitalization", "Oxford", "Php, Pat", "CMS-1500"),
        ("Insurance", "Individual Therapy", "UBH", "Cathy, Cam", "CMS-1500"),
    ])
    # WM is decided by the Program Level column, so set it on Wes's row.
    program_level_col = HEADERS.index("Program Level") + 1
    ws.cell(3, program_level_col).value = "OP WM"

    assign_staff(ws, WEDNESDAY, assign_cathy=True)
    staff = _staff_by_client(ws)

    assert staff["Self, Sam"] == "CB"        # Self Pay is always CB's
    assert staff["Wm, Wes"] == "Melissa"     # only Melissa bills WM
    assert staff["Php, Pat"] == "Melissa"    # PHP is always Melissa's
    assert staff["Cathy, Cam"] == "Cathy"


def test_cathy_takes_iop_for_her_payers():
    """Every Professional service for Cathy's payers is hers, IOP included."""
    ws = _sheet([
        ("Insurance", "IOP", "Oxford", "Iop, Ida", "CMS-1500"),
        ("Insurance", "Telemed IOP", "ConnectiCare", "Iop, Ivan", "UB-04"),
        ("Insurance", "Detox Admission", "UBH", "Detox, Dora", "CMS-1500"),
        ("Insurance", "E-Care Individual", "Oxford", "Ecare, Ellis", "CMS-1500"),
        # IOP for a payer Cathy does not cover is still Jasmine's.
        ("Insurance", "Telemed IOP", "Optum", "Iop, Otto", "CMS-1500"),
        # IOP for one of her payers, but not a Professional claim type, is
        # still Jasmine's.
        ("Insurance", "IOP", "Oxford", "Iop, Inst", "837I"),
    ])
    assign_staff(ws, WEDNESDAY, assign_cathy=True)
    staff = _staff_by_client(ws)

    assert staff["Iop, Ida"] == "Cathy"
    assert staff["Iop, Ivan"] == "Cathy"
    # Professional bills every day, so Wednesday Detox/e-care are hers too.
    assert staff["Detox, Dora"] == "Cathy"
    assert staff["Ecare, Ellis"] == "Cathy"

    assert staff["Iop, Otto"] == "Jasmine"
    assert staff["Iop, Inst"] == "Jasmine"

    # Her rows are labelled "Cathy" in the Staff/Status column, so the
    # Masters workbook names her as the owner too.
    assignments = [ws.cell(row, 1).value for row in range(2, ws.max_row + 1)]
    assert assignments.count("Cathy") == 4


def _cathy_all_payer_rows():
    """Professional rows for the payers only Cathy's full list covers."""
    return [
        ("Insurance", "Individual Therapy", "Emblem (Optum)", "Gold, Gil", "CMS-1500"),
        ("Insurance", "Individual Therapy", "Surest (Optum)", "Hall, Hana", "UB-04"),
        ("Insurance", "Individual Therapy", "UMR (Optum)", "Ives, Ike", "CMS-1500"),
        ("Insurance", "Individual Therapy", "UBH-HP (Optum)", "Jones, Jo", "CMS-1500"),
    ]


def test_cathy_all_payers_adds_the_rest_of_her_list():
    """The full payer list gives Cathy Emblem, Surest, and UMR rows too."""
    rows = _cathy_candidate_rows() + _cathy_all_payer_rows()

    # Her usual three payers only: the added payers stay in the pool.
    ws = _sheet(rows)
    assign_staff(ws, WEDNESDAY, assign_cathy=True)
    staff = _staff_by_client(ws)
    assert staff["Gold, Gil"] == "Rosanna"
    assert staff["Hall, Hana"] == "Rosanna"
    assert staff["Ives, Ike"] == "Rosanna"
    # UBH-HP matches the UBH pattern, so it is hers on either list.
    assert staff["Jones, Jo"] == "Cathy"

    # Full payer list: all four are hers, on top of her usual three.
    ws = _sheet(rows)
    assign_staff(ws, WEDNESDAY, assign_cathy=True, cathy_all_payers=True)
    staff = _staff_by_client(ws)
    for client in ("Gold, Gil", "Hall, Hana", "Ives, Ike", "Jones, Jo",
                   "Adams, Ann", "Baker, Bob", "Carter, Cal"):
        assert staff[client] == "Cathy", (client, staff[client])

    # Still Professional-only, and still nothing outside her payer list.
    assert staff["Diaz, Dee"] == "Unable to Bill"   # 837I Oxford
    assert staff["Evans, Eve"] == "Rosanna"         # Aetna
    assert staff["Frank, Fay"] == "Rosanna"         # plain Optum


def test_cathy_all_payers_turns_the_report_on_by_itself():
    """cathy_all_payers alone runs the Cathy report; assign_cathy is not needed."""
    ws = _sheet(_cathy_candidate_rows() + _cathy_all_payer_rows())
    assign_staff(ws, WEDNESDAY, cathy_all_payers=True)
    staff = _staff_by_client(ws)

    assert staff["Adams, Ann"] == "Cathy"
    assert staff["Gold, Gil"] == "Cathy"


def test_cathy_all_payers_rows_leave_the_professional_pool():
    """Her wider payer list shrinks the pool Rosanna's cap is applied to."""
    rows = [("Insurance", "Individual Therapy", "Emblem (Optum)", f"Cathy{i:04d}",
             "CMS-1500") for i in range(40)]
    rows += [("Insurance", "Individual Therapy", "Magellan", f"Pool{i:04d}",
              "CMS-1500") for i in range(160)]

    ws = _sheet(rows)
    assign_staff(ws, WEDNESDAY, cathy_all_payers=True)
    assignments = [ws.cell(row, 1).value for row in range(2, ws.max_row + 1)]

    assert assignments.count("Cathy") == 40
    # 160 rows are left in the pool: Rosanna's 150, then Jasmine's 10.
    assert assignments.count("Rosanna") == 150
    assert assignments.count("Jasmine") == 10
    # 200 rows in, 200 rows out, each with exactly one owner.
    assert len(assignments) == 200
    assert all(value for value in assignments)


def test_skip_rosanna_gives_the_whole_pool_to_jasmine():
    """With skip_rosanna, Rosanna gets nothing and Jasmine takes the pool."""
    rows = [("Insurance", "Individual Therapy", "Magellan", f"Pool{i:04d}",
             "CMS-1500") for i in range(200)]
    rows.append(("Self Pay", "Individual Therapy", "Self Pay", "Self, Sam", "CMS-1500"))

    # Without the option, Wednesday's cap gives Rosanna her 150.
    ws = _sheet(rows)
    assign_staff(ws, WEDNESDAY)
    assignments = [ws.cell(row, 1).value for row in range(2, ws.max_row + 1)]
    assert assignments.count("Rosanna") == 150
    assert assignments.count("Jasmine") == 50

    ws = _sheet(rows)
    assign_staff(ws, WEDNESDAY, skip_rosanna=True)
    assignments = [ws.cell(row, 1).value for row in range(2, ws.max_row + 1)]

    assert "Rosanna" not in assignments
    assert assignments.count("Jasmine") == 200
    # Nothing is left unassigned, and Self Pay still goes to CB.
    assert assignments.count("CB") == 1
    assert len(assignments) == 201
    assert all(value for value in assignments)


def test_skip_rosanna_with_cathy_on_her_full_payer_list():
    """The two options combine: Cathy takes her payers, Jasmine takes the rest."""
    rows = _cathy_candidate_rows() + _cathy_all_payer_rows()
    ws = _sheet(rows)
    assign_staff(ws, WEDNESDAY, cathy_all_payers=True, skip_rosanna=True)
    staff = _staff_by_client(ws)

    assert "Rosanna" not in staff.values()
    assert staff["Adams, Ann"] == "Cathy"
    assert staff["Gold, Gil"] == "Cathy"
    # Professional rows for payers neither option covers fall to Jasmine.
    assert staff["Evans, Eve"] == "Jasmine"
    assert staff["Frank, Fay"] == "Jasmine"


def test_programming_included_on_a_wednesday():
    """include_programming makes Wednesday Detox/Residential billable to Jasmine."""
    rows = [
        ("Insurance", "Detox Admission", "Oxford", "Detox, Dan", "837I"),
        ("Insurance", "Residential Program", "Optum", "Res, Rita", "837I"),
        ("Insurance", "E-Care Individual", "Optum", "Ecare, Ed", "837I"),
    ]

    ws = _sheet(rows)
    assign_staff(ws, WEDNESDAY)
    staff = _staff_by_client(ws)
    assert staff["Detox, Dan"] == "Unable to Bill"
    assert staff["Res, Rita"] == "Unable to Bill"

    ws = _sheet(rows)
    assign_staff(ws, WEDNESDAY, include_programming=True)
    staff = _staff_by_client(ws)
    assert staff["Detox, Dan"] == "Jasmine"
    assert staff["Res, Rita"] == "Jasmine"
    # e-care is untouched by the override: still Tuesday-only.
    assert staff["Ecare, Ed"] == "Unable to Bill"

    ws = _sheet(rows)
    assign_staff(ws, TUESDAY, include_programming=True)
    staff = _staff_by_client(ws)
    assert staff["Ecare, Ed"] == "Jasmine"


def test_aetna_programming_still_goes_to_melissa():
    """Including Programming does not move Aetna/Humana Detox off Melissa."""
    ws = _sheet([
        ("Insurance", "Detox Admission", "Aetna", "Aetna, Amy", "837I"),
        ("Insurance", "Residential Program", "Humana", "Humana, Hal", "837I"),
    ])
    assign_staff(ws, WEDNESDAY, include_programming=True, assign_cathy=True)
    staff = _staff_by_client(ws)

    assert staff["Aetna, Amy"] == "Melissa"
    assert staff["Humana, Hal"] == "Melissa"


def test_rosanna_cap_applies_to_the_pool_left_after_cathy():
    """Cathy's rows are removed before Rosanna's 150-row cap is applied."""
    rows = [("Insurance", "Individual Therapy", "Oxford", f"Cathy{i:04d}", "CMS-1500")
            for i in range(40)]
    rows += [("Insurance", "Individual Therapy", "Optum", f"Pool{i:04d}", "CMS-1500")
             for i in range(160)]

    ws = _sheet(rows)
    assign_staff(ws, WEDNESDAY, assign_cathy=True)
    assignments = [ws.cell(row, 1).value for row in range(2, ws.max_row + 1)]

    assert assignments.count("Cathy") == 40
    # 160 rows are left in the pool: Rosanna's 150, then Jasmine's 10.
    assert assignments.count("Rosanna") == 150
    assert assignments.count("Jasmine") == 10


def _dropdown_options(**kwargs):
    """Run finalize_workbook on a small workbook and read back its Status list."""
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    ws.append(["Rosanna", "OP Westchester", "Insurance", "Individual Therapy",
               "Oxford", "Smith, John", "OP", "Adams, Ann", "CMS-1500"])
    unbilled.finalize_workbook(wb, **kwargs)
    sheet2 = wb["Sheet2"]
    return [sheet2.cell(row, 1).value for row in range(1, sheet2.max_row + 1)]


def test_cathy_status_dropdown_matches_jasmine():
    """Cathy's Status dropdown carries the same options as Jasmine's."""
    base = ["Billed", "Unable to Bill", "Contractual Adj", "Incomplete Billings",
            "Utox Batch", "Inclusive Services"]

    # Rosanna's list: the six shared options.
    assert _dropdown_options() == base

    # Jasmine's and Cathy's list: the same six plus Batch Billings and IOP.
    jasmine = _dropdown_options(include_batch_billings=True, include_iop_status=True)
    assert jasmine == base + ["Batch Billings", "IOP"]


def test_rosanna_cap_override_replaces_standard_schedule():
    """rosanna_cap_override replaces the weekday schedule with an exact count."""
    rows = [("Insurance", "Individual Therapy", "Magellan", f"Pool{i:04d}",
             "CMS-1500") for i in range(200)]

    ws = _sheet(rows)
    assign_staff(ws, WEDNESDAY, rosanna_cap_override=75)
    assignments = [ws.cell(row, 1).value for row in range(2, ws.max_row + 1)]

    assert assignments.count("Rosanna") == 75
    assert assignments.count("Jasmine") == 125


def test_rosanna_cap_override_can_open_a_weekend():
    """The override also applies on a weekend, which otherwise caps Rosanna at 0."""
    saturday = "09052026"  # a Saturday
    rows = [("Insurance", "Individual Therapy", "Magellan", f"Pool{i:04d}",
             "CMS-1500") for i in range(50)]

    ws = _sheet(rows)
    assign_staff(ws, saturday)
    assignments = [ws.cell(row, 1).value for row in range(2, ws.max_row + 1)]
    assert "Rosanna" not in assignments

    ws = _sheet(rows)
    assign_staff(ws, saturday, rosanna_cap_override=20)
    assignments = [ws.cell(row, 1).value for row in range(2, ws.max_row + 1)]
    assert assignments.count("Rosanna") == 20
    assert assignments.count("Jasmine") == 30


def test_rosanna_cap_override_ignored_when_skip_rosanna():
    """skip_rosanna wins over rosanna_cap_override: Rosanna still gets nothing."""
    rows = [("Insurance", "Individual Therapy", "Magellan", f"Pool{i:04d}",
             "CMS-1500") for i in range(50)]

    ws = _sheet(rows)
    assign_staff(ws, WEDNESDAY, rosanna_cap_override=20, skip_rosanna=True)
    assignments = [ws.cell(row, 1).value for row in range(2, ws.max_row + 1)]

    assert "Rosanna" not in assignments
    assert assignments.count("Jasmine") == 50


def test_custom_report_routes_matching_professional_payer_rows():
    """A custom report claims Professional rows for its payers, like a second Cathy."""
    ws = _sheet(_cathy_candidate_rows())
    assign_staff(ws, WEDNESDAY, custom_report_name="Karen",
                 custom_report_payer_terms=["aetna"])
    staff = _staff_by_client(ws)

    assert staff["Evans, Eve"] == "Karen"  # Aetna, CMS-1500
    # Optum wasn't in the custom report's payer list, so it stays Rosanna's.
    assert staff["Frank, Fay"] == "Rosanna"
    # Oxford isn't in this custom report's payer list ("aetna"), so this row
    # is untouched by it either way; it's a non-Professional claim type
    # outside the daily schedule, so it's Unable to Bill regardless.
    assert staff["Diaz, Dee"] == "Unable to Bill"


def test_custom_report_any_claim_type_when_professional_only_is_false():
    """custom_report_professional_only=False matches any claim type, not just Professional."""
    rows = [
        ("Insurance", "Individual Therapy", "Cigna", "Adams, Ann", "CMS-1500"),
        ("Insurance", "Individual Therapy", "Cigna", "Baker, Bob", "837I"),
    ]
    ws = _sheet(rows)
    assign_staff(ws, WEDNESDAY, custom_report_name="Karen",
                 custom_report_payer_terms=["cigna"],
                 custom_report_professional_only=False)
    staff = _staff_by_client(ws)

    assert staff["Adams, Ann"] == "Karen"
    assert staff["Baker, Bob"] == "Karen"


def test_custom_report_leaves_pool_for_rosanna_and_jasmine():
    """Custom report rows leave the professional pool entirely, same as Cathy's."""
    rows = [("Insurance", "Individual Therapy", "Cigna", f"Karen{i:04d}", "CMS-1500")
            for i in range(40)]
    rows += [("Insurance", "Individual Therapy", "Optum", f"Pool{i:04d}", "CMS-1500")
             for i in range(160)]

    ws = _sheet(rows)
    assign_staff(ws, WEDNESDAY, custom_report_name="Karen",
                 custom_report_payer_terms=["cigna"])
    assignments = [ws.cell(row, 1).value for row in range(2, ws.max_row + 1)]

    assert assignments.count("Karen") == 40
    assert assignments.count("Rosanna") == 150
    assert assignments.count("Jasmine") == 10


def test_custom_report_and_cathy_do_not_double_claim():
    """When both are on, Cathy is checked first; the custom report never re-claims her rows."""
    rows = [
        ("Insurance", "Individual Therapy", "Oxford", "Adams, Ann", "CMS-1500"),
        ("Insurance", "Individual Therapy", "Cigna", "Baker, Bob", "CMS-1500"),
    ]
    ws = _sheet(rows)
    # A custom report configured to also match Oxford: Cathy still gets it,
    # because assign_cathy is checked first in assign_staff.
    assign_staff(ws, WEDNESDAY, assign_cathy=True,
                 custom_report_name="Karen",
                 custom_report_payer_terms=["oxford", "cigna"])
    staff = _staff_by_client(ws)

    assert staff["Adams, Ann"] == "Cathy"
    assert staff["Baker, Bob"] == "Karen"


def test_custom_report_inactive_without_both_name_and_payers():
    """Setting only the name or only the payer terms leaves the standard schedule in place."""
    rows = [("Insurance", "Individual Therapy", "Cigna", "Adams, Ann", "CMS-1500")]

    ws = _sheet(rows)
    assign_staff(ws, WEDNESDAY, custom_report_name="Karen")
    assert _staff_by_client(ws)["Adams, Ann"] == "Rosanna"

    ws = _sheet(rows)
    assign_staff(ws, WEDNESDAY, custom_report_payer_terms=["cigna"])
    assert _staff_by_client(ws)["Adams, Ann"] == "Rosanna"


def test_validate_custom_report_name_rejects_reserved_names():
    """Reserved staff names (case-insensitive) are rejected for the custom report."""
    for name in ("Rosanna", "jasmine", "CB", "Melissa", "cathy", "Unable to Bill"):
        try:
            unbilled.validate_custom_report_name(name)
            assert False, f"expected ValueError for reserved name {name!r}"
        except ValueError:
            pass

    # A non-reserved name, and no name at all, are both fine.
    unbilled.validate_custom_report_name("Karen")
    unbilled.validate_custom_report_name(None)
    unbilled.validate_custom_report_name("")


if __name__ == '__main__':
    for name, test in sorted(
        (name, obj) for name, obj in list(globals().items())
        if name.startswith('test_') and callable(obj)
    ):
        test()
        print(f"✓ {name} passed")

    print("\nAll tests passed!")
