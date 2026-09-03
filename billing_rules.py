"""
Helper functions to:
- parse MMDDYYYY date tokens into weekday
- determine whether a given service is non-billable for a given weekday
- classify a "Claim Type" value as Professional (CMS-1500 or UB-04)

Weekday mapping: 0=Monday, 1=Tuesday, ..., 6=Sunday

Daily billing schedule:
- Professional services (Claim Type "CMS-1500" or "UB-04" — UB-04 counts as
  Professional every day of the week) bill every day of the week.
- IOP (including Telemed IOP) bills every day of the week, with no
  exceptions.
- Programming (Detox, Residential) bills Tuesday, Thursday, Friday, and
  weekends; non-billable Monday and Wednesday.
- E-care bills on Tuesdays only, regardless of Claim Type.
- Self Pay: every service bills every day, with no exceptions (including
  e-care).
- PHP (Partial Hospitalization) is not part of the Programming schedule
  above: it always goes to Melissa (see is_php_service), and is billed
  only on Tuesdays as a real-world/manual matter, not something this
  module's weekday helpers gate.

Rosanna/Jasmine split (applies to GROUPFLD2 == "Insurance" rows only; no
Self Pay ever goes to either of them, and PHP rows never reach them since
PHP always goes to Melissa):
- The "professional pool" is every Insurance row whose Claim Type is
  CMS-1500 or UB-04, sorted alphabetically by Client. Monday through
  Friday, Rosanna receives the first ROSANNA_PROFESSIONAL_CAP[weekday]
  rows (150) of that sorted pool; anything past her share of the pool
  goes to Jasmine. Rosanna caps no rows on weekends, so the entire pool
  goes to Jasmine those days.
- Jasmine also receives every billable Programming/e-care row, plus any
  other billable Insurance row whose Claim Type is not CMS-1500/UB-04
  (i.e. institutional/837I), unless it's PHP (always Melissa's).
- IOP (including Telemed IOP) always goes to Jasmine, every day of the
  week, bypassing the professional pool/Rosanna split entirely, even if
  Claim Type is CMS-1500 or UB-04.
- GROUPFLD2 values other than "Insurance" or "Self Pay" never reach
  Rosanna or Jasmine.

Optional, per-run overrides (all off by default; nothing below changes the
standard schedule unless the operator turns it on for that run):
- include_programming: bill Programming (Detox, Residential) regardless of
  the weekday, so it can be included on a Monday or Wednesday. E-care is
  unaffected and stays Tuesday-only.
- Cathy report: pull every Professional (CMS-1500/UB-04) Insurance row
  whose Payer is Oxford, ConnectiCare, or UBH (see is_cathy_payer) out of
  the professional pool and assign it to Cathy instead, so a row is never
  worked twice. Whatever the service is, it is hers — including IOP for
  those three payers, which she takes ahead of the IOP-to-Jasmine rule.
  WM, PHP and the O'Flynn Karen rule still take priority over it.
- Cathy report, all of her payers: the same report run against Cathy's
  full payer list (see is_cathy_all_payer / CATHY_ALL_PAYERS) instead of
  just her usual three. Only the payer list widens; it is still
  Professional Insurance rows only, and it turns the Cathy report on by
  itself.
- Nothing for Rosanna: give Rosanna no rows at all for the run. Her share
  of the professional pool goes to Jasmine instead, exactly as it does on
  a weekend, and no workbook is generated for her.
- Exclude Aetna: drop Aetna rows (see is_aetna_payer) from the individual
  staff reports; they stay in the Masters workbook.
"""
import re
from datetime import datetime
from typing import Tuple

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


DRUG_SCREEN_KEYWORDS = ("drug screen", "utox", "urine tox", "drug test", "uds")
PHP_KEYWORDS = ("partial hospitalization", "php")

# Rosanna's professional-service row cap by weekday (0=Monday..6=Sunday).
# Rosanna works Monday through Friday, capped at 150 rows/day; she receives
# no rows on weekends.
ROSANNA_PROFESSIONAL_CAP = {0: 150, 1: 150, 2: 150, 3: 150, 4: 150}

# Payers that go to Cathy's Professional-services-only report when that
# optional report is turned on. Matching is case-insensitive and tolerant of
# the spelling variants these payers appear with in the Payer column
# ("ConnectiCare"/"Connecti Care", "UBH"/"United Behavioral Health").
CATHY_PAYERS = ("Oxford", "ConnectiCare", "UBH")
_CATHY_PAYER_RE = re.compile(
    r"oxford|connecti[\s\-]?care|\bubh\b|united\s+behavioral\s+health",
    re.IGNORECASE,
)

# The full payer list Cathy can be given for a run: her usual three plus the
# Optum-family plans that sit alongside them in the report's payer filter.
# "UBH-HP" is spelled out here for the operator's benefit, but it already
# matches the UBH pattern above, so it is Cathy's under either payer list.
CATHY_ALL_PAYERS = ("ConnectiCare", "Emblem", "Oxford", "Surest", "UBH",
                    "UBH-HP", "UMR")
_CATHY_ALL_PAYER_RE = re.compile(
    r"oxford|connecti[\s\-]?care|\bubh\b|united\s+behavioral\s+health"
    r"|emblem|surest|\bumr\b",
    re.IGNORECASE,
)


def parse_weekday_from_token(date_token: str) -> Tuple[int, bool]:
    """
    Parse an MMDDYYYY date_token and return (weekday, did_fallback).
    - weekday: integer 0..6
    - did_fallback: True if parsing failed and we fell back to today's weekday
    """
    if not date_token:
        return datetime.now().weekday(), True
    try:
        dt = datetime.strptime(date_token, "%m%d%Y")
        return dt.weekday(), False
    except Exception:
        return datetime.now().weekday(), True


def _is_ecare(service: str) -> bool:
    """Return True if the service text refers to e-care (many possible variants)."""
    s = service.lower()
    return any(token in s for token in ("e-care", "e care", "ecare", "extended care"))


def _is_drug_screen(service: str) -> bool:
    """Return True if the service text refers to a drug screen (Utox)."""
    s = service.lower()
    return any(kw in s for kw in DRUG_SCREEN_KEYWORDS)


def is_php_service(service: str) -> bool:
    """Return True for PHP/Partial Hospitalization services (always Melissa's)."""
    s = (service or "").lower()
    return any(kw in s for kw in PHP_KEYWORDS)


def _is_programming_service(service: str) -> bool:
    """Return True for Programming services: Detox, Residential.

    IOP is intentionally excluded here: it bills every day of the week
    (see is_iop_service) rather than following this weekday schedule.
    PHP is also excluded: it's always assigned to Melissa (see
    is_php_service) rather than following this weekday schedule.
    """
    s = service.lower()
    return (
        "detox" in s
        or "residential" in s
    )


def is_iop_service(service: str) -> bool:
    """Return True for any IOP service, including Telemed IOP (case-insensitive)."""
    return "iop" in (service or "").lower()


def is_professional_claim_type(claim_type: str) -> bool:
    """Return True if the Claim Type column value is CMS-1500 or UB-04 (Professional).

    UB-04 counts as Professional every day of the week, alongside CMS-1500.
    """
    return (claim_type or "").strip().upper() in ("CMS-1500", "UB-04")


def is_cathy_payer(payer: str) -> bool:
    """Return True if the Payer column value is Oxford, ConnectiCare, or UBH.

    Used by the optional Cathy report, which takes only Professional
    (CMS-1500/UB-04) Insurance rows for these three payers.
    """
    return bool(_CATHY_PAYER_RE.search(payer or ""))


def is_cathy_all_payer(payer: str) -> bool:
    """Return True if the Payer column value is on Cathy's full payer list.

    A superset of is_cathy_payer: her usual three payers plus Emblem,
    Surest, UBH-HP and UMR. Used by the optional "all of her payers"
    variant of the Cathy report, which is otherwise identical — still
    Professional (CMS-1500/UB-04) Insurance rows only.
    """
    return bool(_CATHY_ALL_PAYER_RE.search(payer or ""))


def is_aetna_payer(payer: str) -> bool:
    """Return True if the Payer column value refers to Aetna."""
    return "aetna" in (payer or "").lower()


def parse_terms(raw: str) -> list:
    """Split a comma-separated string into trimmed, non-empty terms.

    Used by the custom per-run payer/service exclusion fields, so an
    operator can type "Cigna, Humana" into a text box instead of a new
    checkbox needing a code change for each new payer or service.
    """
    if not raw:
        return []
    return [term.strip() for term in raw.split(",") if term.strip()]


def matches_any_term(text: str, terms) -> bool:
    """Return True if text contains any of terms, case-insensitively."""
    if not terms:
        return False
    haystack = (text or "").lower()
    return any(term.lower() in haystack for term in terms)


def is_wm_program_level(cell_value) -> bool:
    """Return True if the Program Level cell contains 'WM' (covers OP WM too)."""
    return "WM" in str(cell_value or "").upper()


def is_anthem_payer(payer: str) -> bool:
    """Return True if the payer contains 'anthem'."""
    return "anthem" in (payer or "").lower()


def is_bcb_anthem_ct_php_res_detox(payer: str, service: str) -> bool:
    """Return True if payer is BCB Anthem CT and service is PHP, Residential, or Detox."""
    return ("bcb anthem ct" in (payer or "").lower() and
            any(s in (service or "").lower() for s in ["partial hospitalization", "residential", "detox"]))


# Staff names assign_staff already routes rows to on its own. A custom
# report's name (see assign_staff's custom_report_name) must not collide
# with one of these, or its rows would be indistinguishable from that
# staff's own.
RESERVED_STAFF_NAMES = ("Rosanna", "Jasmine", "CB", "Melissa", "Cathy", "Unable to Bill")


def validate_custom_report_name(name: str) -> None:
    """Raise ValueError if name collides with a reserved staff name."""
    if name and name.strip().lower() in {n.lower() for n in RESERVED_STAFF_NAMES}:
        raise ValueError(
            f"'{name}' is already a reserved staff name "
            f"({', '.join(RESERVED_STAFF_NAMES)}); choose a different name "
            "for the custom report."
        )


def assign_staff(ws, date_token: str = None, include_programming: bool = False,
                 assign_cathy: bool = False, cathy_all_payers: bool = False,
                 skip_rosanna: bool = False, rosanna_cap_override: int = None,
                 custom_report_name: str = None, custom_report_payer_terms: list = None,
                 custom_report_professional_only: bool = True):
    """Assign staff names based on the standard daily billing rules.

    This is the single copy of the assignment engine: app.py (the Streamlit
    app) and "Unbilled Step 1.py" (the CLI script) both import it from here
    instead of each keeping their own copy, so a rule change can no longer
    happen in one and not the other.

    - Self Pay (GROUPFLD2 == "Self Pay") always goes to CB; every service
      bills every day, with no exceptions.
    - Jasmine and Rosanna only ever receive GROUPFLD2 == "Insurance" rows.
    - Among Insurance rows, the "professional pool" is every row whose
      Claim Type is CMS-1500 or UB-04 (UB-04 counts as Professional every
      day), sorted alphabetically by Client. Monday through Friday,
      Rosanna receives the first ROSANNA_PROFESSIONAL_CAP[weekday] (150)
      of that sorted pool; the rest of the pool goes to Jasmine. Rosanna
      caps no rows on weekends, so Jasmine gets the whole pool those days.
    - Jasmine also receives Insurance rows that are billable Programming
      (Detox, Residential) or e-care for that weekday, as well as any
      other billable Insurance row whose Claim Type is not CMS-1500/UB-04
      (i.e. institutional/837I), unless it's PHP (always Melissa's).
    - IOP (including Telemed IOP) always goes to Jasmine, every day of the
      week, bypassing the professional pool/Rosanna split even if Claim
      Type is CMS-1500 or UB-04 — unless the Cathy report or a custom
      report claims the row first (see below).
    - Any other Insurance row (not billable that day), or any row that is
      neither Self Pay nor Insurance, is Unable to Bill.
    - Melissa (WM/OP WM Program Level, PHP/Partial Hospitalization, or
      Aetna/Humana Detox/Residential) and the O'Flynn Karen OP
      Chappaqua/OP NYC "Unable to Bill" rule take priority over all of the
      above. PHP rows are assigned to Melissa every day in the Masters
      spreadsheet; she does not get an individual report, and PHP is
      billed only on Tuesdays as an operational matter.

    Optional, per-run overrides (all off by default) sit on top of the
    schedule above:
    - include_programming: Programming (Detox, Residential) is billable
      regardless of the weekday, so it can be worked on a Monday or
      Wednesday. E-care is unaffected and stays Tuesday-only.
    - assign_cathy: every Professional (CMS-1500/UB-04) Insurance row whose
      Payer is Oxford, ConnectiCare, or UBH goes to Cathy instead of into
      the Rosanna/Jasmine professional pool, so no row is worked twice.
      Whatever the service is, it is hers: that includes IOP for those
      three payers, which she takes ahead of the IOP-to-Jasmine rule. The
      rules ahead of her (Self Pay, WM, O'Flynn Karen, Aetna/Humana
      Detox/Residential, and PHP) still take priority.
    - cathy_all_payers: run that same Cathy rule against her full payer
      list (CATHY_ALL_PAYERS) instead of just her usual three — adding
      Emblem, Surest, UBH-HP and UMR. Only the payer list widens: it is
      still Professional Insurance rows only, and the same rules still
      take priority over her. This turns the Cathy report on by itself,
      whether or not assign_cathy is also set.
    - skip_rosanna: Rosanna is given no rows at all. The whole professional
      pool left after Cathy's/the custom report's goes to Jasmine, exactly
      as it does on a weekend.
    - rosanna_cap_override: give Rosanna exactly this many professional-pool
      rows for this run instead of the standard weekday schedule. Ignored
      (Rosanna gets zero) when skip_rosanna is also set.
    - custom_report_name/custom_report_payer_terms/
      custom_report_professional_only: a second, generic "Cathy slot" for
      routing a specific payer's rows to a different named staff member
      without a code change. When custom_report_name and
      custom_report_payer_terms are both set, every Insurance row whose
      Payer contains one of those terms is that staff's — same placement
      as Cathy (ahead of the IOP-to-Jasmine rule and the professional
      pool), checked after Cathy so the two never claim the same row.
      custom_report_professional_only (default True, matching Cathy)
      restricts it to CMS-1500/UB-04 claim types; set it False to match
      any claim type instead. The name must not collide with a reserved
      staff name (see RESERVED_STAFF_NAMES/validate_custom_report_name).

    Args:
        ws: Worksheet to process
        date_token: Date string in MMDDYYYY format (from filename). If None, uses current date.
        include_programming: When True, bill Programming regardless of weekday.
        assign_cathy: When True, route Oxford/ConnectiCare/UBH Professional
            Insurance rows to Cathy.
        cathy_all_payers: When True, give Cathy her full payer list instead
            of her usual three, and turn her report on by itself.
        skip_rosanna: When True, assign Rosanna nothing; Jasmine takes the
            whole professional pool.
        rosanna_cap_override: When set, use this row count as Rosanna's cap
            for this run instead of the standard weekday schedule.
        custom_report_name: When set (with custom_report_payer_terms), the
            staff name to assign matching rows to.
        custom_report_payer_terms: When set (with custom_report_name), payer
            terms (case-insensitive substring match) that route a row to
            custom_report_name.
        custom_report_professional_only: When True (default), the custom
            report only claims Professional (CMS-1500/UB-04) rows, like
            Cathy. Set False to match any claim type.
    """

    # Find column indices (after Staff/Status insert, columns shift by 1)
    cols = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        if header == "GROUPFLD1":
            cols['group_fld1'] = col
        elif header == "GROUPFLD2":
            cols['group'] = col
        elif header == "Service":
            cols['service'] = col
        elif header == "Payer":
            cols['payer'] = col
        elif header == "Billing Provider":
            cols['billing_provider'] = col
        elif header == "Program Level":
            cols['program_level'] = col
        elif header == "Client":
            cols['client'] = col
        elif header == "Claim Type":
            cols['claim_type'] = col

    if not all(k in cols for k in ['group', 'service', 'payer', 'claim_type', 'client']):
        raise ValueError(
            "Missing required columns: GROUPFLD2, Service, Payer, Claim Type, or Client"
        )

    weekday, did_fallback = parse_weekday_from_token(date_token)
    if did_fallback:
        print(f"Failed to parse date_token '{date_token}', using current weekday={weekday}")
    else:
        print(f"Using date from filename: {date_token} (weekday={weekday})")

    # "All of her payers" implies the Cathy report: checking that option
    # alone is enough to run it.
    assign_cathy = assign_cathy or cathy_all_payers
    is_cathy_row_payer = is_cathy_all_payer if cathy_all_payers else is_cathy_payer

    custom_report_active = bool(custom_report_name and custom_report_payer_terms)

    # Rosanna takes nothing when skip_rosanna is on, so the whole pool falls
    # to Jasmine — the same path a weekend already takes. Otherwise, an
    # explicit per-run override wins over the standard weekday schedule.
    if skip_rosanna:
        rosanna_cap = 0
    elif rosanna_cap_override is not None:
        rosanna_cap = max(0, rosanna_cap_override)
    else:
        rosanna_cap = ROSANNA_PROFESSIONAL_CAP.get(weekday, 0)
    if rosanna_cap:
        capped_staff, professional_cap = "Rosanna", rosanna_cap
    else:
        capped_staff, professional_cap = None, 0

    row_data_map = {}
    fixed_staff = {}        # original_row -> staff already decided
    professional_rows = []  # (original_row, client) still needing the capped-staff/Jasmine split
    other_rows = []         # original_row order for every row not in the professional pool

    for row in range(2, ws.max_row + 1):
        row_data_map[row] = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]

        group = str(ws.cell(row, cols['group']).value or "").strip()
        service = str(ws.cell(row, cols['service']).value or "")
        payer = str(ws.cell(row, cols['payer']).value or "").lower()
        claim_type = str(ws.cell(row, cols['claim_type']).value or "")

        # CB: Self Pay bills every service every day, no exceptions. This
        # overrides every other rule.
        if group == "Self Pay":
            fixed_staff[row] = "CB"
            other_rows.append(row)
            continue

        staff = None

        # WM / OP WM Program Level → Melissa
        if 'program_level' in cols and is_wm_program_level(ws.cell(row, cols['program_level']).value):
            staff = "Melissa"

        # Unable to Bill: Billing Provider = "O'Flynn, Karen" + GROUPFLD1 = "OP Chappaqua" or "OP NYC"
        if not staff and 'billing_provider' in cols and 'group_fld1' in cols:
            billing_provider = str(ws.cell(row, cols['billing_provider']).value or "").strip()
            group_fld1 = str(ws.cell(row, cols['group_fld1']).value or "").strip()
            if (billing_provider == "O'Flynn, Karen" and
                    group_fld1 in ("OP Chappaqua", "OP NYC")):
                staff = "Unable to Bill"

        # Melissa: (Detox or Residential) + (Aetna or Humana), but not drug screens
        if not staff:
            service_lower = service.lower()
            has_detox_res = ("detox" in service_lower or "residential" in service_lower)
            has_insurance = "aetna" in payer or "humana" in payer
            if has_detox_res and has_insurance and not _is_drug_screen(service):
                staff = "Melissa"

        # PHP/Partial Hospitalization always goes to Melissa, every day.
        if not staff and is_php_service(service):
            staff = "Melissa"

        if staff:
            fixed_staff[row] = staff
            other_rows.append(row)
            continue

        # Jasmine and Rosanna only ever receive Insurance rows.
        if group != "Insurance":
            fixed_staff[row] = "Unable to Bill"
            other_rows.append(row)
            continue

        # Cathy (optional): every Professional row for her payers is hers,
        # whatever the service is. That includes IOP, so this sits ahead of
        # the IOP-to-Jasmine rule below. Her rows leave the Rosanna/Jasmine
        # professional pool entirely rather than being worked twice.
        if (assign_cathy and is_professional_claim_type(claim_type)
                and is_cathy_row_payer(payer)):
            fixed_staff[row] = "Cathy"
            other_rows.append(row)
            continue

        # Custom report (optional): a second, generic Cathy-shaped slot for
        # routing a specific payer's rows to a different staff member,
        # checked after Cathy so the two never claim the same row. Same
        # placement as Cathy — ahead of the IOP-to-Jasmine rule and the
        # professional pool.
        if (custom_report_active
                and (not custom_report_professional_only or is_professional_claim_type(claim_type))
                and matches_any_term(payer, custom_report_payer_terms)):
            fixed_staff[row] = custom_report_name
            other_rows.append(row)
            continue

        # IOP (including Telemed IOP) always goes to Jasmine, every day of
        # the week, bypassing the professional pool/Rosanna split even if
        # Claim Type is CMS-1500 or UB-04.
        if is_iop_service(service):
            fixed_staff[row] = "Jasmine"
            other_rows.append(row)
            continue

        if is_professional_claim_type(claim_type):
            client = str(ws.cell(row, cols['client']).value or "").strip()
            professional_rows.append((row, client))
            continue

        if is_non_billable_service_for_weekday(
                service, weekday, include_programming=include_programming):
            fixed_staff[row] = "Unable to Bill"
        else:
            fixed_staff[row] = "Jasmine"
        other_rows.append(row)

    # Move the professional (Insurance + CMS-1500/UB-04) pool to the top of
    # the sheet, sorted alphabetically by Client; every other row keeps its
    # original relative order after that.
    professional_rows.sort(key=lambda x: x[1].lower())

    ordered_rows = [row for row, _ in professional_rows] + other_rows
    new_row_pos = 2
    for original_row in ordered_rows:
        for col in range(1, ws.max_column + 1):
            ws.cell(new_row_pos, col).value = row_data_map[original_row][col - 1]
        new_row_pos += 1

    num_capped = min(professional_cap, len(professional_rows))
    new_row_pos = 2
    for idx in range(len(professional_rows)):
        ws.cell(new_row_pos, 1).value = capped_staff if (capped_staff and idx < num_capped) else "Jasmine"
        new_row_pos += 1
    for original_row in other_rows:
        ws.cell(new_row_pos, 1).value = fixed_staff[original_row]
        new_row_pos += 1

    print("Staff assignment complete")


def finalize_workbook(wb, include_batch_billings: bool = False, include_iop_status: bool = False,
                      skip_status_columns: bool = False):
    """Add Status/Comments columns and validation for Rosanna/Jasmine/Cathy/custom exports.

    Args:
        wb: Workbook to finalize.
        include_batch_billings: When True, add 'Batch Billings' to the dropdown
            (Jasmine, Cathy, and custom reports).
        include_iop_status: When True, add 'IOP' to the dropdown (Jasmine,
            Cathy, and custom reports).
        skip_status_columns: When True, skip adding Status/Comments columns (for CB/self-pay).
    """
    ws = wb.active

    # Bold header row
    for col in range(1, ws.max_column + 1):
        ws.cell(1, col).font = openpyxl.styles.Font(bold=True)

    # Insert two columns at E (unless skipped for CB/self-pay)
    if not skip_status_columns:
        ws.insert_cols(5, 2)
        ws.cell(1, 5).value = "Status"
        ws.cell(1, 6).value = "Comments"
        ws.cell(1, 5).font = openpyxl.styles.Font(bold=True)
        ws.cell(1, 6).font = openpyxl.styles.Font(bold=True)

        # Create Sheet2 with validation list
        ws_list = wb.create_sheet("Sheet2")
        status_items = ["Billed", "Unable to Bill", "Contractual Adj", "Incomplete Billings",
                        "Utox Batch", "Inclusive Services"]
        if include_batch_billings:
            status_items.append("Batch Billings")
        if include_iop_status:
            status_items.append("IOP")

        for idx, item in enumerate(status_items, start=1):
            ws_list[f'A{idx}'] = item
        list_range = f"=Sheet2!$A$1:$A${len(status_items)}"

        # Add data validation to Status column
        dv = DataValidation(type="list", formula1=list_range, allow_blank=True)
        ws.add_data_validation(dv)

        last_row = ws.max_row
        dv.add(f'E2:E{last_row}')

    # Widen columns to fit all text. openpyxl's auto_size flag is unreliable in
    # Excel, so compute an explicit width from the longest value in each column.
    for col in range(1, ws.max_column + 1):
        max_length = 0
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row, col).value
            if value is None:
                continue
            length = len(str(value))
            if length > max_length:
                max_length = length
        # Add a small padding and cap the width so a single huge cell doesn't
        # blow out the layout.
        ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 60)

    # Enable filtering on the header row.
    ws.auto_filter.ref = ws.dimensions


def is_non_billable_service_for_weekday(
    service: str,
    weekday: int,
    is_professional: bool = False,
    self_pay: bool = False,
    include_programming: bool = False,
) -> bool:
    """
    Return True if the given service (string) should be treated as non-billable
    on the specified weekday.

    - service: original service string (case-insensitive checks will be used)
    - weekday: 0=Monday .. 6=Sunday
    - is_professional: True when the row's Claim Type is CMS-1500 or UB-04.
      Professional rows bill every day of the week.
    - self_pay: when True, every service is billable every day, with no
      exceptions (including e-care).
    - include_programming: one-off override. When True, Programming (Detox,
      Residential) is billable regardless of the weekday, so it can be
      included on a Monday or Wednesday. It does not affect e-care, which
      stays Tuesday-only.
    """
    s = (service or "").lower().strip()

    # Self Pay bills every service every day, no exceptions.
    if self_pay:
        return False

    # Professional (CMS-1500) services bill every day of the week.
    if is_professional:
        return False

    # e-care is only billed on Tuesdays.
    if _is_ecare(s):
        return weekday != 1

    # IOP (including Telemed IOP) bills every day of the week.
    if is_iop_service(s):
        return False

    # Programming (Detox, Residential) bills Tue/Thu/Fri + weekends, unless
    # the one-off include_programming override is on for this run.
    if _is_programming_service(s):
        if include_programming:
            return False
        return weekday not in (1, 3, 4, 5, 6)

    # Anything else is neither Professional nor Programming nor e-care, and
    # isn't addressed by the daily schedule, so it is not billable.
    return True
